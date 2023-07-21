###############################################################################################################
#    weatherData.py    Copyright (C) <2023>  <Kevin Scott>                                                    #
#                                                                                                             #
#    A class that loads in a spread sheet containing weather data.                                            #
#                                                                                                             #
###############################################################################################################
#    Copyright (C) <2023>  <Kevin Scott>                                                                      #
#                                                                                                             #
#    This program is free software: you can redistribute it and/or modify it under the terms of the           #
#    GNU General Public License as published by the Free Software Foundation, either Version 3 of the         #
#    License, or (at your option) any later Version.                                                          #
#                                                                                                             #
#    This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without        #
#    even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the               #
#    GNU General Public License for more details.                                                             #
#                                                                                                             #
#    You should have received a copy of the GNU General Public License along with this program.               #
#    If not, see <http://www.gnu.org/licenses/>.                                                              #
#                                                                                                             #
###############################################################################################################

import json
from json import JSONEncoder

import colorama

from openpyxl import load_workbook, Workbook

import src.classes.dataMapping as DM

from src.classes.dataClasses import weatherValues


class WeatherData():
    """  A class that loads in a spread sheet containing weather data.

         mainData = WD.WeatherData("data\\main.xlsx")  -  pass in the file containing the data in excell spreadsheet format.

         The file is loaded and the data inserted into a dictionary, then the object is created.

         mainData.loadFile()   -  loads the spreadsheet.
         mainData.loadData()   -  loads the spreadsheet into a directory.
         mainData.saveData()   -  save the directory back to the spreadsheet.
         mainData.countData()  -  return the number of entries.
         mainData.nextRow()    -  returns the data, one line at a time.
         mainData.add()        -  adds a row into the directory.

         to see in an entry is in the directory use key in mainData (uses __contains__)

    """

    def __init__(self, file, screen=True):
        if screen:
            print(f"{colorama.Fore.GREEN} Processing data file {file}{colorama.Fore.RESET}")

        self.screen = screen
        self.fileName = file
        self.values   = {}         #  Dictionary to hold weather values.
        self.val_keys = ()         #  List to hold the dictionary keys
        self.position = 0
        self.loadFile()
        self.loadData()


    def __contains__(self, key):
        """  checks to see if key already exists in the data directory.
             The directory keys are help in a separate list and MUST be kept updated.
        """
        return key in self.val_keys


    def loadFile(self):
        """  Load the spreadsheet only load data and not formulas.
        """
        try:
            self.workBook = load_workbook(filename=self.fileName, read_only=False, data_only=True)
        except FileNotFoundError:
            print(f"{self.fileName} not found, create blank file.")
            self.workBook = Workbook()      #  Initial spreadsheet dosn't exist - create a new blank one.

        #  Grab the active worksheet
        self.workSheet = self.workBook.active
        if self.screen:
            print(f"{colorama.Fore.GREEN}     Scanning {self.workSheet.title} in {self.fileName}{colorama.Fore.RESET}")

    def loadData(self):
        """  Loads the entries into the directory.
        """
        for rowNumber in range(DM.START_ROW, self.workSheet.max_row):

            self.wv = weatherValues(
                self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_TEMPERATURE).value,
                self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_FEELS_LIKE).value,
                self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_DEW_POINT).value,
                self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_HUMIDITY).value,
                self.workSheet.cell(row=rowNumber, column=DM.INDOOR_TEMPERATURE).value,
                self.workSheet.cell(row=rowNumber, column=DM.INDOOR_HUMIDITY).value,
                self.workSheet.cell(row=rowNumber, column=DM.SOLAR).value,
                self.workSheet.cell(row=rowNumber, column=DM.UVI).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_RATE).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_DAILY).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_EVENT).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_HOURLY).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_WEEKLY).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_MONTHLY).value,
                self.workSheet.cell(row=rowNumber, column=DM.RAIN_YEARLY).value,
                self.workSheet.cell(row=rowNumber, column=DM.WIND_SPEED).value,
                self.workSheet.cell(row=rowNumber, column=DM.WIND_GUST).value,
                self.workSheet.cell(row=rowNumber, column=DM.WIND_DIRECTION).value,
                self.workSheet.cell(row=rowNumber, column=DM.PRESSURE_RELATIVE).value,
                self.workSheet.cell(row=rowNumber, column=DM.PRESSURE_ABSOLUTE).value
            )  #  A class to hold each row of weather data.  **REMEMBER**

            self.key                   = self.workSheet.cell(row=rowNumber, column=DM.TIME).value

            self.values[self.key]      = self.wv

            self.wv = None      #  Is this needed? - memory??

        self.val_keys = list(self.values.keys())    #  A list of keys for later.


    def countData(self):
        """  Return the number of entries.
        """
        return len(self.values)


    def saveDataJson(self, jsonFileName="dump.json"):
        """  Converts the data directory to a json object, then writes to a file.
             Needs a custom encoder (dataEncoder), weatherValues is not natively serializable.
        """
        json_object = json.dumps(self.values, indent=4, cls=dataEncoder)
        with open(jsonFileName, "w") as json_file:
             json_file.write(json_object)


    def nextRow(self):
        """  Returns the data, one line at a time.
        """
        if self.position <= self.countData():
            self.position += 1
            yield (self.val_keys[self.position], self.values[self.val_keys[self.position]])     #  Return a tuple (key, data)


    def add(self, key, row):
        """  Adds a new entry to the data dictionary.
             The key is added to the list of keys.
        """
        self.values[key] = row
        self.val_keys.append(key)


    def saveData(self):
        """  Save the data directory to a excel spreadsheet.
        """
        rowNumber = DM.START_ROW
        for key, row in self.values.items():

            # the .cell is a function, we ignore the return value - looks like the cell reference.
            self.workSheet.cell(row=rowNumber, column=DM.TIME               , value = key)
            self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_TEMPERATURE, value = row.OutdoorTemperature )
            self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_FEELS_LIKE , value = row.OutdoorFeelsLike)
            self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_DEW_POINT  , value = row.OutdoorDewPoint)
            self.workSheet.cell(row=rowNumber, column=DM.OUTDOOR_HUMIDITY   , value = row.OutdoorHumidity)
            self.workSheet.cell(row=rowNumber, column=DM.INDOOR_TEMPERATURE , value = row.IndoorTemperature)
            self.workSheet.cell(row=rowNumber, column=DM.INDOOR_HUMIDITY    , value = row.IndoorHumidity)
            self.workSheet.cell(row=rowNumber, column=DM.SOLAR              , value = row.Solar)
            self.workSheet.cell(row=rowNumber, column=DM.UVI                , value = row.UVI)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_RATE          , value = row.RainRate)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_DAILY         , value = row.RainDaily)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_EVENT         , value = row.RainEvent)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_HOURLY        , value = row.RainHourly)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_WEEKLY        , value = row.RainWeekly)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_MONTHLY       , value = row.RainMonthly)
            self.workSheet.cell(row=rowNumber, column=DM.RAIN_YEARLY        , value = row.RainYearly)
            self.workSheet.cell(row=rowNumber, column=DM.WIND_SPEED         , value = row.WindSpeed)
            self.workSheet.cell(row=rowNumber, column=DM.WIND_GUST          , value = row.WindGust)
            self.workSheet.cell(row=rowNumber, column=DM.WIND_DIRECTION     , value = row.WindDirection)
            self.workSheet.cell(row=rowNumber, column=DM.PRESSURE_RELATIVE  , value = row.PressureRelative)
            self.workSheet.cell(row=rowNumber, column=DM.PRESSURE_ABSOLUTE  , value = row.PressureAbsolute)

            rowNumber += 1

        self.workBook.save(filename=self.fileName)



class dataEncoder(JSONEncoder):
    """  A custom json encoder to bake directories serializable.
    """
    def default(self, o):
        return o.__dict__





